"""
@File_name:    create_performance_case/create_case
@Author:         liuwei
@Time:            2023/2/10 21:36
"""
import datetime
import os
import re
import shutil
from openpyxl import load_workbook


class CreateCase(object):
    def __init__(self, source_file, target_file=None):
        self.source_file = source_file
        self.target_file = target_file

    @staticmethod
    def open_excel(file_name, sheet_name=None):
        path = os.getcwd() + "/excel/" + file_name
        excel = load_workbook(filename=path)
        if "工况性能" in excel.sheetnames:
            sheet = excel["工况性能"]
        else:
            sheet = excel[sheet_name]
        excel.close()
        return sheet

    @staticmethod
    def copy_excel(source_file, destination_file):
        current_path = os.getcwd()
        source_file_path = current_path + "/excel/case/" + source_file
        destination_file_path = current_path + "/excel/case/" + destination_file
        shutil.copy(source_file_path, destination_file_path)

    def get_result_from_excel(self):
        result_dict = {}
        sheet = CreateCase.open_excel(self.source_file)
        rows_list = []
        for row in sheet.rows:
            row = list(map(lambda x: x.value, row))
            rows_list.append(row)
        for i in range(len(rows_list) - 1):
            if "倍率" in rows_list[i]:
                setattr(CreateCase, "rate_list", rows_list[i][rows_list[i].index("倍率") + 1::])
            elif "最大幅度" in rows_list[i]:
                setattr(CreateCase, "arm_list", rows_list[i][rows_list[i].index("最大幅度") + 1::])
            elif rows_list[i][0] is None or "Steps" in rows_list[i]:
                if isinstance(rows_list[i][1], float):
                    x, y = str(rows_list[i][1]).split(".")
                    setattr(CreateCase, f"rated_weight_list_{x}_point_{y}",
                            list(filter(bool, rows_list[i][2::])))
                else:
                    setattr(CreateCase, f"rated_weight_list_{rows_list[i][1]}",
                            list(filter(bool, rows_list[i][2::])))

        arm_rate_list = []
        arm_list = getattr(CreateCase, "arm_list")
        for i in range(len(arm_list)):
            if i % 2 == 0:
                arm_rate_list.append(str(arm_list[i]) + "_2")
            else:
                arm_rate_list.append(str(arm_list[i]) + "_4")

        for attr in dir(CreateCase):
            if "rated_weight_list_" in attr:
                scope = attr.split("list_")[-1]
                arm_rate_scope_list = list(map(lambda x: x + "_" + str(scope), arm_rate_list))
                attr_value_list = getattr(CreateCase, attr)
                result_dict.update(dict(zip(arm_rate_scope_list, attr_value_list)))
        return result_dict

    def modify_excel(self, file_name, data):
        path = os.getcwd() + "/excel/" + file_name
        excel = load_workbook(filename=path)
        sheet = excel["工况性能"]
        for row_num in range(1, sheet.max_row + 1):
            for column_num in range(1, sheet.max_column):
                old_value = sheet.cell(row_num, column_num).value
                if old_value is None:
                    continue
                else:
                    if "$" in str(old_value):
                        regx = "\\${(.+?)}"
                        key = re.search(regx, old_value).group(1)
                        new_str = re.sub(regx, str(data.get(key)), old_value)
                        sheet.cell(row_num, column_num).value = new_str
        excel.save(path)
        excel.close()


if __name__ == '__main__':
    now_str = datetime.datetime.now().strftime("%Y-%d-%m~%H-%M-%S")
    destination_file = f"{now_str}.xlsx"
    CreateCase.copy_excel(source_file="case_template.xlsx", destination_file=destination_file)
    create_case = CreateCase("performance/performance.xlsx", target_file=destination_file)
    data = create_case.get_result_from_excel()
    create_case.modify_excel(f"case/{destination_file}", data)
